import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

# Load the CSV file into a DataFrame
df = pd.read_csv('CSE.csv')
df = df.fillna(method='ffill')

# Rename columns to match your headers
df.rename(columns={
    'Student': 'student_name',
    'Registration Id': 'registration_id',
    'Batch Year': 'batch_year',
    'Programme Section': 'programme_section',
    'Course': 'course_name',
    'Present': 'attendance_percentage',
    'Overall Present': 'overall_attendance'
}, inplace=True)


# Exclude courses with "PRACTICAL" or "LECTURE" tags
df = df[~df['course_name'].str.contains('PRACTICAL|LECTURE', case=False, na=False)]

# List of project courses
project_courses = ['Basic Internship [CSE(INT)301]  [PROJECT]', 'Intermediate Internship [CSE(INT)501]  [PROJECT]', 'Advanced Internship [CSE(INT)701]  [PROJECT]']

# Apply the eligibility logic
def determine_eligibility(row, project_in_courses):
    if project_in_courses:
        return 'Allowed for all courses due to Project'
    elif row['overall_attendance'] >= 75:
        return 'Allowed for all courses'
    elif row['attendance_percentage'] >= 75:
        return 'Allowed for this course'
    else:
        return 'Not allowed'

# Create a new DataFrame to store eligibility data
eligibility_data = pd.DataFrame()

# Check if any student has one of the "Project" courses
students = df['student_name'].unique()

for student in students:
    student_data = df[df['student_name'] == student].copy()
    project_in_courses = student_data['course_name'].isin(project_courses).any()
    student_data.loc[:, 'eligibility'] = student_data.apply(determine_eligibility, axis=1, project_in_courses=project_in_courses)
    eligibility_data = pd.concat([eligibility_data, student_data])

# Sort the DataFrame by student_name and course_name
df = eligibility_data[['student_name', 'registration_id', 'batch_year', 'programme_section', 'course_name', 'attendance_percentage', 'overall_attendance', 'eligibility']].sort_values(by=['student_name', 'course_name'])

# Create an Excel workbook and add a worksheet
wb = Workbook()
ws = wb.active
ws.title = "Student Exam Eligibility"

# Add the DataFrame data to the worksheet
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# Merge cells for student_name, overall_attendance, and eligibility where applicable
current_student = None
start_row = 2  # Excel rows start from 1, and the first row is the header
project_allowed = False  # Flag to track if "Allowed for all courses due to Project" condition is met

for row in range(2, ws.max_row + 1):
    student_name = ws.cell(row=row, column=1).value
    eligibility = ws.cell(row=row, column=8).value
    overall_attendance = ws.cell(row=row, column=7).value
    
    # Track if "Allowed for all courses due to Project" condition is met
    if eligibility == 'Allowed for all courses due to Project':
        project_allowed = True
    
    if student_name != current_student:
        if current_student is not None:
            # Merge student_name cells
            ws.merge_cells(start_row=start_row, start_column=1, end_row=row-1, end_column=1)
            # Merge eligibility cells for "Allowed for all courses", "Allowed for all courses due to Project", or "Not allowed"
            if ws.cell(row=start_row, column=8).value == 'Allowed for all courses' or project_allowed or \
               ws.cell(row=start_row, column=8).value == 'Not allowed':
                ws.merge_cells(start_row=start_row, start_column=8, end_row=row-1, end_column=8)
            # Merge overall_attendance cells
            ws.merge_cells(start_row=start_row, start_column=7, end_row=row-1, end_column=7)
        current_student = student_name
        start_row = row
        project_allowed = eligibility == 'Allowed for all courses due to Project'  # Reset for new student

# Handle the last student group
if current_student is not None:
    ws.merge_cells(start_row=start_row, start_column=1, end_row=ws.max_row, end_column=1)
    if ws.cell(row=start_row, column=8).value == 'Allowed for all courses' or project_allowed or \
       ws.cell(row=start_row, column=8).value == 'Not allowed':
        ws.merge_cells(start_row=start_row, start_column=8, end_row=ws.max_row, end_column=8)
    # Merge overall_attendance cells
    ws.merge_cells(start_row=start_row, start_column=7, end_row=ws.max_row, end_column=7)

# Center align the merged cells
for col in ['A', 'G', 'H']:
    for cell in ws[col]:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Save the workbook
wb.save('student_exam_eligibility.xlsx')